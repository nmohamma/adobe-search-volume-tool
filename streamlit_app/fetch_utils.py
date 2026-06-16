"""
Fetch Adobe Stock Video search counts via Jina AI Reader.
Standalone module for Streamlit app — user provides their own API key.
"""
import urllib.parse, re, time
import requests


def fetch_count(search_term: str, api_key: str, delay: float = 2.5) -> str | None:
    """
    Fetch live Adobe Stock Video search count for a quoted search term.

    Args:
        search_term: The term to search (e.g. "New Year's Day")
        api_key: Jina AI API key
        delay: Seconds to wait before request (rate limiting)

    Returns:
        Formatted count string (e.g. "796,544") or None on failure.
    """
    if delay > 0:
        time.sleep(delay)

    quoted = f'"{search_term}"'
    url = f"https://stock.adobe.com/search/video?k={urllib.parse.quote(quoted)}"
    jina_url = f"https://r.jina.ai/{url}"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/plain",
        "Authorization": f"Bearer {api_key}",
    }

    try:
        resp = requests.get(jina_url, headers=headers, timeout=30)
        resp.raise_for_status()
        text = resp.text

        patterns = [
            r'\*\*([0-9,]+)\s+results?\s+for\s+',
            r'([0-9,]+)\s+results?\s+for\s+',
            r'([0-9,]+)\s+results',
        ]
        for pat in patterns:
            m = re.search(pat, text)
            if m:
                raw = m.group(1).replace(",", "")
                # Filter absurd parsing artifacts (>100M)
                try:
                    if int(raw) > 100_000_000:
                        return None
                except ValueError:
                    pass
                return m.group(1)
        return None
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 429:
            return None  # rate limited, silently skip
        return None
    except Exception:
        return None


def parse_events(text: str) -> list[dict]:
    """Parse pasted text into structured events.
    
    Handles:
    - One event per line (newline‑separated)
    - Multiple events on the same line separated by commas
    - Optional date prefix (e.g., "January 1 - New Year's Day")
    
    Returns a list of dictionaries with keys "date" and "name".
    """
    events = []
    date_pattern = re.compile(r'^([A-Za-z]+\s+\d+(?:-\d+)?\s*(?:\([^)]*\))?)\s*[-\–—]\s*(.*)')
    
    for raw_line in text.strip().split('\n'):
        line = raw_line.strip()
        if not line:
            continue
        # If the line matches a date prefix, treat the whole line as a single event
        m = date_pattern.match(line)
        if m:
            events.append({"date": m.group(1).strip(), "name": m.group(2).strip()})
            continue
        # Split on commas
        parts = [p.strip() for p in line.split(',')]
        for part in parts:
            if part:
                events.append({"date": "", "name": part})
    return events


def build_xlsx(events: list[dict], term_map: dict[str, str | None]) -> bytes:
    """Build XLSX from events + counts, return as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Search Volumes"

    headers = ["Date", "Event Name", "Search Count", "Count (Numeric)"]
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9E2F3'),
        right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'),
        bottom=Side(style='thin', color='D9E2F3'),
    )
    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = [22, 50, 18, 15][col - 1]

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    alt = False
    for i, ev in enumerate(events):
        r = i + 2
        count = term_map.get(ev["name"])
        count_display = f"{int(count):,}" if count and count.isdigit() else ""
        count_num = int(count.replace(",", "")) if count and count.replace(",", "").isdigit() else 0

        vals = [ev["date"], ev["name"], count_display, count_num]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name='Calibri', size=10, color='006600' if col == 3 and count_display else None)
            c.border = thin_border
            if col >= 3:
                c.alignment = Alignment(horizontal='right')
            if alt:
                c.fill = alt_fill
        alt = not alt

    ws.auto_filter.ref = f"A1:D{len(events) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
